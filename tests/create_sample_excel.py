"""テスト用サンプルExcelファイルを生成するスクリプト"""

import openpyxl
import os

wb = openpyxl.Workbook()

# Sheet 1: 月次売上データ
ws1 = wb.active
ws1.title = "月次売上"
ws1.append(["月", "売上（万円）", "前年比（%）", "目標達成率（%）"])
ws1.append(["4月", 1200, 105, 96])
ws1.append(["5月", 1350, 112, 108])
ws1.append(["6月", 980, 89, 78])
ws1.append(["7月", 1500, 125, 120])
ws1.append(["8月", 1100, 95, 88])
ws1.append(["9月", 1420, 118, 114])

# Sheet 2: 地域別売上
ws2 = wb.create_sheet("地域別売上")
ws2.append(["地域", "売上（万円）", "構成比（%）", "前期比（%）"])
ws2.append(["東京", 3200, 35, 110])
ws2.append(["大阪", 2100, 23, 105])
ws2.append(["名古屋", 1500, 16, 98])
ws2.append(["福岡", 1200, 13, 115])
ws2.append(["札幌", 800, 9, 92])
ws2.append(["その他", 400, 4, 88])

# Sheet 3: 製品カテゴリ
ws3 = wb.create_sheet("製品カテゴリ")
ws3.append(["カテゴリ", "売上（万円）", "利益率（%）"])
ws3.append(["ソフトウェア", 4500, 45])
ws3.append(["ハードウェア", 2800, 15])
ws3.append(["サービス", 1900, 35])

output_dir = os.path.dirname(os.path.abspath(__file__))
filepath = os.path.join(output_dir, "fixtures", "sample.xlsx")
os.makedirs(os.path.dirname(filepath), exist_ok=True)
wb.save(filepath)
print(f"Created: {filepath}")
