"""Excelファイルの読み込みとテーブル検出"""

import openpyxl
from openpyxl.utils import get_column_letter
from modules.models import ParsedTable


def load_workbook(file) -> openpyxl.Workbook:
    """ファイルオブジェクトまたはパスからワークブックを読み込む"""
    return openpyxl.load_workbook(file, data_only=True)


def get_sheet_names(wb: openpyxl.Workbook) -> list[str]:
    """シート名一覧を返す"""
    return wb.sheetnames


def detect_tables(ws, sheet_name: str) -> list[ParsedTable]:
    """シート内のテーブルを自動検出する。
    ヒューリスティック: 連続する非空セルの矩形領域をテーブルとして認識。
    """
    tables = []
    if ws.max_row is None or ws.max_column is None:
        return tables

    visited = set()
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 50)):
            if (row, col) in visited:
                continue
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue

            # テーブルの範囲を検出
            table = _extract_table_from(ws, sheet_name, row, col, visited)
            if table and len(table.headers) >= 2 and len(table.rows) >= 1:
                tables.append(table)

    return tables


def _extract_table_from(ws, sheet_name: str, start_row: int, start_col: int, visited: set) -> ParsedTable | None:
    """指定位置からテーブルの範囲を検出して抽出する"""
    # 右方向に拡張してヘッダー列数を決定
    end_col = start_col
    for col in range(start_col, min(ws.max_column + 1, start_col + 50)):
        if ws.cell(row=start_row, column=col).value is not None:
            end_col = col
        else:
            break

    # 下方向に拡張して行数を決定
    end_row = start_row
    for row in range(start_row + 1, min(ws.max_row + 1, start_row + 500)):
        # 行の中に1つでも値があれば続行
        has_value = False
        for col in range(start_col, end_col + 1):
            if ws.cell(row=row, column=col).value is not None:
                has_value = True
                break
        if has_value:
            end_row = row
        else:
            break

    # 訪問済みとしてマーク
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            visited.add((r, c))

    # ヘッダー行（最初の行）を取得
    headers = []
    for col in range(start_col, end_col + 1):
        val = ws.cell(row=start_row, column=col).value
        headers.append(str(val) if val is not None else "")

    # データ行を取得
    rows = []
    for row in range(start_row + 1, end_row + 1):
        row_data = []
        for col in range(start_col, end_col + 1):
            val = ws.cell(row=row, column=col).value
            row_data.append(val)
        rows.append(row_data)

    # 範囲文字列を生成
    start_ref = f"{get_column_letter(start_col)}{start_row}"
    end_ref = f"{get_column_letter(end_col)}{end_row}"
    source_range = f"{sheet_name}!{start_ref}:{end_ref}"

    # タイトルはヘッダーの最初の値を使用（仮）
    title = headers[0] if headers else None

    return ParsedTable(
        sheet_name=sheet_name,
        title=title,
        headers=headers,
        rows=rows,
        source_range=source_range,
    )


def parse_range(ws, sheet_name: str, range_str: str) -> ParsedTable:
    """ユーザー指定の範囲からテーブルを抽出する。
    range_str: "A1:D10" 形式
    """
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(range_str)

    headers = []
    for col in range(min_col, max_col + 1):
        val = ws.cell(row=min_row, column=col).value
        headers.append(str(val) if val is not None else "")

    rows = []
    for row in range(min_row + 1, max_row + 1):
        row_data = []
        for col in range(min_col, max_col + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        rows.append(row_data)

    source_range = f"{sheet_name}!{range_str}"
    return ParsedTable(
        sheet_name=sheet_name,
        title=headers[0] if headers else None,
        headers=headers,
        rows=rows,
        source_range=source_range,
    )
